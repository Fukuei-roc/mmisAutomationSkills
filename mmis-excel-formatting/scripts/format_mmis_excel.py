from __future__ import annotations

import argparse
from copy import copy, deepcopy
import json
import logging
import re
import sys
import unicodedata
import warnings
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


TARGET_DIR = Path(
    r"C:\Users\NMMIS\OneDrive - Ministry of Transportation and Communications-7280502-Taiwan Railways Administration, MOTC\文件\MMIS桌面"
)
DEFAULT_FILE_TYPE = "auto"
SHEET_NAME = "故障通報管理 的清單"
DATE_HEADER = "發生日期"
CAR_HEADER = "車組/車號"
NEW_HEADER = "車號"
FONT_NAME = "新細明體"
FONT_SIZE = 12
FAULT_NOTICE_FILENAME_PATTERNS = (
    "故障通報管理",
    "未處理故障通報",
)
OPEN_FAULT_NOTICE_FILENAME_PATTERNS = ("未結案故障通報",)
COMMON_DELETE_COLUMNS = [
    "發生時間",
    "故障地點",
    "立案人員",
    "通報人員",
    "通報單位",
    "狀態",
    "配屬段別",
    "配屬段別名稱",
]
UNPROCESSED_FAULT_NOTICE_WIDTHS = {
    "車次": 5.7,
    "車組/車號": 11.6,
    "車號": 5.7,
    "發生日期": 10.8,
    "事故等級": 10.8,
    "ATP故障": 10.4,
    "故障現象": 70.7,
    "通報號": 11.6,
}
OPEN_FAULT_NOTICE_WIDTHS: dict[str, float] = {
    "車次": 5.7,
    "車組/車號": 11.6,
    "發生日期": 10.8,
    "事故等級": 10.8,
    "ATP故障": 10.4,
    "故障現象": 70.7,
    "負責單位": 10.8,
    "通報號": 11.6,
    "待簽核工單": 14.2,
}

warnings.filterwarnings(
    "ignore",
    message="Workbook contains no default style, apply openpyxl's default",
    module="openpyxl.styles.stylesheet",
)


class FormattingError(RuntimeError):
    pass


def build_logger() -> logging.Logger:
    logger = logging.getLogger("mmis_excel_formatting")
    if logger.handlers:
        return logger

    logger.setLevel(logging.INFO)
    handler = logging.StreamHandler(sys.stderr)
    handler.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
    logger.addHandler(handler)
    return logger


LOGGER = build_logger()


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def list_existing_files() -> list[str]:
    if not TARGET_DIR.exists():
        return []
    return sorted(item.name for item in TARGET_DIR.iterdir() if item.is_file())


def result_template(*, file_path: Path | None = None, file_type: str = DEFAULT_FILE_TYPE) -> dict[str, Any]:
    filename = file_path.name if file_path else None
    path = str(file_path) if file_path else None
    return {
        "file_found": False,
        "filename": filename,
        "file_type": file_type,
        "detected_type": None,
        "applied_config": None,
        "format_applied": False,
        "sorted": False,
        "sorting_applied": [],
        "column_inserted": False,
        "header_set": False,
        "value_range": None,
        "value_verification": [],
        "unparsed_count": 0,
        "saved": False,
        "reason": None,
        "existing_files": [],
        "sheet_names": [],
        "headers": [],
        "path": path,
        "used_range": None,
        "verification": [],
        "full_range_verified": False,
        "mismatch_cells": [],
        "selected_by": None,
        "deleted_headers": [],
        "header_bold_applied": False,
        "layout_applied": False,
        "autofit_applied": False,
        "custom_column_widths_applied": [],
    }


def find_used_range(ws) -> tuple[int, int, int, int]:
    min_row = None
    min_col = None
    max_row = 0
    max_col = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                if min_row is None or cell.row < min_row:
                    min_row = cell.row
                if min_col is None or cell.column < min_col:
                    min_col = cell.column
                if cell.row > max_row:
                    max_row = cell.row
                if cell.column > max_col:
                    max_col = cell.column
    if min_row is None or min_col is None:
        return (1, 1, ws.max_row or 1, ws.max_column or 1)
    return (min_row, min_col, max_row, max_col)


def get_column_index_map(ws) -> dict[str, int]:
    return {
        normalize_text(ws.cell(row=1, column=col_idx).value): col_idx
        for col_idx in range(1, ws.max_column + 1)
        if normalize_text(ws.cell(row=1, column=col_idx).value)
    }


def get_column_indexes_map(ws) -> dict[str, list[int]]:
    column_map: dict[str, list[int]] = {}
    for col_idx in range(1, ws.max_column + 1):
        header = normalize_text(ws.cell(row=1, column=col_idx).value)
        if not header:
            continue
        column_map.setdefault(header, []).append(col_idx)
    return column_map


def getColumnIndexMap(ws) -> dict[str, int]:
    return get_column_index_map(ws)


def remove_columns_by_name(ws, columns: list[str]) -> list[str]:
    column_map = get_column_index_map(ws)
    deletions = [(column_map[column_name], column_name) for column_name in columns if column_name in column_map]
    removed: list[str] = []
    for col_idx, column_name in sorted(deletions, key=lambda item: item[0], reverse=True):
        ws.delete_cols(col_idx, 1)
        removed.append(column_name)
    return list(reversed(removed))


def removeColumnsByName(ws, columns: list[str]) -> list[str]:
    return remove_columns_by_name(ws, columns)


def parse_date_for_sort(value: Any) -> tuple[int, str]:
    if value is None or value == "":
        return (9, "")
    if hasattr(value, "strftime"):
        return (0, value.strftime("%Y%m%d%H%M%S"))
    text = normalize_text(value)
    candidates = [
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%m/%d/%Y",
        "%m-%d-%Y",
    ]
    for fmt in candidates:
        try:
            parsed = datetime.strptime(text, fmt)
            return (0, parsed.strftime("%Y%m%d%H%M%S"))
        except ValueError:
            continue
    digits = "".join(ch for ch in text if ch.isdigit())
    if len(digits) >= 8:
        return (1, digits)
    return (8, text)


def parse_car_for_sort(value: Any) -> tuple[int, list[Any], str]:
    text = normalize_text(value)
    if not text:
        return (9, [], "")
    parts: list[Any] = []
    current = ""
    current_is_digit = text[0].isdigit()
    for ch in text:
        if ch.isdigit() == current_is_digit:
            current += ch
        else:
            parts.append(int(current) if current_is_digit else current.lower())
            current = ch
            current_is_digit = ch.isdigit()
    parts.append(int(current) if current_is_digit else current.lower())
    return (0, parts, text.lower())


def parse_sort_value(value: Any, parser_name: str) -> Any:
    if parser_name == "date":
        return parse_date_for_sort(value)
    if parser_name == "car":
        return parse_car_for_sort(value)
    return (9, "") if value in (None, "") else (0, normalize_text(value).lower())


def sort_worksheet(ws, sort_rules: list[dict[str, str]]) -> list[str]:
    min_row, min_col, max_row, max_col = find_used_range(ws)
    if max_row <= min_row:
        return []

    column_map = get_column_index_map(ws)
    active_rules = [rule for rule in sort_rules if rule["column"] in column_map]
    if not active_rules:
        return []

    data_rows: list[list[Any]] = []
    for row_idx in range(min_row + 1, max_row + 1):
        values = [ws.cell(row=row_idx, column=col_idx).value for col_idx in range(min_col, max_col + 1)]
        data_rows.append(values)

    def sort_key(row: list[Any]) -> tuple[Any, ...]:
        key_parts: list[Any] = []
        for rule in active_rules:
            col_idx = column_map[rule["column"]] - min_col
            key_parts.append(parse_sort_value(row[col_idx], rule.get("parser", "text")))
        return tuple(key_parts)

    data_rows.sort(key=sort_key)

    for row_offset, values in enumerate(data_rows, start=min_row + 1):
        for col_idx, value in enumerate(values, start=min_col):
            ws.cell(row=row_offset, column=col_idx).value = value

    return [rule["column"] for rule in active_rules]


def sortWorksheet(ws, sort_rules: list[dict[str, str]]) -> list[str]:
    return sort_worksheet(ws, sort_rules)


def apply_global_font(ws, font_name: str, font_size: int) -> None:
    min_row, min_col, max_row, max_col = find_used_range(ws)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            new_font = copy(cell.font)
            new_font.name = font_name
            new_font.size = font_size
            cell.font = new_font


def applyGlobalFont(ws, font_name: str, font_size: int) -> None:
    apply_global_font(ws, font_name, font_size)


def apply_global_alignment(ws, horizontal: str, vertical: str) -> None:
    min_row, min_col, max_row, max_col = find_used_range(ws)
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            new_alignment = copy(cell.alignment) if cell.alignment else Alignment()
            new_alignment.horizontal = horizontal
            new_alignment.vertical = vertical
            cell.alignment = new_alignment


def applyGlobalAlignment(ws, horizontal: str, vertical: str) -> None:
    apply_global_alignment(ws, horizontal, vertical)


def apply_header_alignment(ws, horizontal: str, vertical: str) -> None:
    column_map = get_column_index_map(ws)
    if not column_map:
        return
    min_col = min(column_map.values())
    max_col = max(column_map.values())
    header_row = next(ws.iter_rows(min_row=1, max_row=1, min_col=min_col, max_col=max_col), ())
    for cell in header_row:
        new_alignment = copy(cell.alignment) if cell.alignment else Alignment()
        new_alignment.horizontal = horizontal
        new_alignment.vertical = vertical
        cell.alignment = new_alignment


def applyHeaderAlignment(ws, horizontal: str, vertical: str) -> None:
    apply_header_alignment(ws, horizontal, vertical)


def display_width(value: Any) -> int:
    text = normalize_text(value)
    if not text:
        return 0
    width = 0
    for ch in text:
        width += 2 if unicodedata.east_asian_width(ch) in {"F", "W"} else 1
    return width


def autofit_columns(ws) -> None:
    column_map = get_column_index_map(ws)
    if not column_map:
        return
    min_col = min(column_map.values())
    max_col = max(column_map.values())
    for col_idx in range(min_col, max_col + 1):
        max_width = 0
        for row_idx in range(1, ws.max_row + 1):
            max_width = max(max_width, display_width(ws.cell(row=row_idx, column=col_idx).value))
        ws.column_dimensions[get_column_letter(col_idx)].width = max(8, min(max_width + 2, 80))


def autofitColumns(ws) -> None:
    autofit_columns(ws)


def apply_column_widths(ws, width_map: dict[str, float]) -> dict[str, int]:
    column_map = get_column_indexes_map(ws)
    applied: dict[str, int] = {}
    for header, width in width_map.items():
        col_indexes = column_map.get(header, [])
        if not col_indexes:
            continue
        for col_idx in col_indexes:
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        applied[header] = len(col_indexes)
    return applied


def applyColumnWidths(ws, width_map: dict[str, float]) -> dict[str, int]:
    return apply_column_widths(ws, width_map)


def ensure_car_number_column(ws, result: dict[str, Any]) -> None:
    _, _, max_row, _ = find_used_range(ws)
    if normalize_text(ws["C1"].value) != NEW_HEADER:
        ws.insert_cols(3, 1)
        result["column_inserted"] = True
    else:
        result["column_inserted"] = True

    ws["C1"] = NEW_HEADER
    result["header_set"] = normalize_text(ws["C1"].value) == NEW_HEADER

    if max_row < 2:
        raise FormattingError("資料列為 0")

    if not any(normalize_text(ws.cell(row=row_idx, column=2).value) for row_idx in range(2, max_row + 1)):
        raise FormattingError("找不到 B 欄")

    unparsed_count = 0
    for row_idx in range(2, max_row + 1):
        source_value = normalize_text(ws.cell(row=row_idx, column=2).value)
        match = re.search(r"(\d+)$", source_value)
        if not match:
            ws.cell(row=row_idx, column=3, value=None)
            unparsed_count += 1
            continue

        n = int(match.group(1))
        result_value = n // 10 if 1000 <= n < 10000 else n
        ws.cell(row=row_idx, column=3, value=result_value)

    result["value_range"] = f"C2~C{max_row}"
    result["unparsed_count"] = unparsed_count


def apply_fault_notice_header_style(ws, result: dict[str, Any]) -> None:
    cell = ws["C1"]
    new_font = copy(cell.font) if cell.font else Font()
    new_font.bold = True
    cell.font = new_font
    result["header_bold_applied"] = bool(ws["C1"].font.bold)


def apply_header_cell_style(ws, cell_ref: str, text: str) -> None:
    cell = ws[cell_ref]
    cell.value = text
    new_font = copy(cell.font) if cell.font else Font()
    new_font.name = FONT_NAME
    new_font.size = FONT_SIZE
    new_font.bold = True
    cell.font = new_font

    new_alignment = copy(cell.alignment) if cell.alignment else Alignment()
    new_alignment.horizontal = "center"
    new_alignment.vertical = "center"
    cell.alignment = new_alignment


def fault_notice_before_format(ws, result: dict[str, Any], config: dict[str, Any], file_path: Path) -> None:
    ensure_car_number_column(ws, result)


def fault_notice_after_format(ws, result: dict[str, Any], config: dict[str, Any], file_path: Path) -> None:
    apply_fault_notice_header_style(ws, result)


def open_fault_notice_after_format(ws, result: dict[str, Any], config: dict[str, Any], file_path: Path) -> None:
    if normalize_text(ws["G1"].value) != "負責單位":
        ws.insert_cols(7, 1)
        LOGGER.info("inserted column at G")

    apply_global_font(ws, config["font_name"], config["font_size"])
    alignment = config["alignment"]
    header_alignment = config["header_alignment"]
    apply_global_alignment(ws, alignment["horizontal"], alignment["vertical"])
    apply_header_alignment(ws, header_alignment["horizontal"], header_alignment["vertical"])

    apply_header_cell_style(ws, "G1", "負責單位")
    LOGGER.info("set header G1 = 負責單位")

    for cell_ref in ("I1", "J1", "K1"):
        apply_header_cell_style(ws, cell_ref, "待簽核工單")
    LOGGER.info("set header I1/J1/K1 = 待簽核工單")

    apply_global_font(ws, config["font_name"], config["font_size"])
    apply_global_alignment(ws, alignment["horizontal"], alignment["vertical"])
    apply_header_alignment(ws, header_alignment["horizontal"], header_alignment["vertical"])
    apply_header_cell_style(ws, "G1", "負責單位")
    for cell_ref in ("I1", "J1", "K1"):
        apply_header_cell_style(ws, cell_ref, "待簽核工單")

    width_map = config.get("post_hook_column_widths", {})
    if width_map:
        LOGGER.info("applying custom column widths")
        applied_counts = apply_column_widths(ws, width_map)
        result["custom_column_widths_applied"] = []
        for header, width in width_map.items():
            count = applied_counts.get(header, 0)
            if count <= 0:
                continue
            if count == 1:
                LOGGER.info("set width: %s = %s", header, width)
            else:
                LOGGER.info("set width: %s (%s columns) = %s", header, count, width)
            result["custom_column_widths_applied"].extend([header] * count)


def get_format_config(file_type: str) -> dict[str, Any]:
    configs = {
        "fault_notice": {
            "name": "FAULT_NOTICE_CONFIG",
            "sheet_name": SHEET_NAME,
            "delete_columns": COMMON_DELETE_COLUMNS,
            "sort_rules": [
                {"column": DATE_HEADER, "parser": "date"},
                {"column": CAR_HEADER, "parser": "car"},
            ],
            "font_name": FONT_NAME,
            "font_size": FONT_SIZE,
            "alignment": {"horizontal": "left", "vertical": "top"},
            "header_alignment": {"horizontal": "center", "vertical": "center"},
            "autofit": True,
            "column_widths": {},
            "before_format": fault_notice_before_format,
            "after_format": fault_notice_after_format,
        },
        "open_fault_notice": {
            "name": "OPEN_FAULT_NOTICE_CONFIG",
            "sheet_name": SHEET_NAME,
            "delete_columns": COMMON_DELETE_COLUMNS,
            "sort_rules": [
                {"column": "事故等級", "parser": "text"},
                {"column": DATE_HEADER, "parser": "date"},
                {"column": CAR_HEADER, "parser": "car"},
            ],
            "font_name": FONT_NAME,
            "font_size": FONT_SIZE,
            "alignment": {"horizontal": "left", "vertical": "top"},
            "header_alignment": {"horizontal": "center", "vertical": "center"},
            "autofit": True,
            "column_widths": {},
            "post_hook_column_widths": OPEN_FAULT_NOTICE_WIDTHS,
            "before_format": None,
            "after_format": open_fault_notice_after_format,
        },
    }
    if file_type not in configs:
        raise FormattingError("尚未支援的檔案類型")
    return deepcopy(configs[file_type])


def getFormatConfig(file_type: str) -> dict[str, Any]:
    return get_format_config(file_type)


def detect_file_type(file_name: str | Path) -> str | None:
    filename = file_name.name if isinstance(file_name, Path) else str(file_name)
    if any(pattern in filename for pattern in OPEN_FAULT_NOTICE_FILENAME_PATTERNS):
        return "open_fault_notice"
    if any(pattern in filename for pattern in FAULT_NOTICE_FILENAME_PATTERNS):
        return "fault_notice"
    return None


def detectFileType(file_name: str | Path) -> str | None:
    return detect_file_type(file_name)


def build_runtime_config(file_type: str, file_path: Path) -> dict[str, Any]:
    config = get_format_config(file_type)
    if file_type == "fault_notice" and "未處理故障通報" in file_path.name:
        config["column_widths"] = deepcopy(UNPROCESSED_FAULT_NOTICE_WIDTHS)
    return config


def apply_formatting(ws, config: dict[str, Any], result: dict[str, Any]) -> None:
    removed_columns = remove_columns_by_name(ws, config.get("delete_columns", []))
    result["deleted_headers"] = removed_columns
    LOGGER.info("removed columns: %s", ", ".join(removed_columns) if removed_columns else "(none)")

    sorting_applied = sort_worksheet(ws, config.get("sort_rules", []))
    result["sorted"] = True
    result["sorting_applied"] = sorting_applied
    LOGGER.info("sorting applied: %s", " -> ".join(sorting_applied) if sorting_applied else "(none)")

    apply_global_font(ws, config["font_name"], config["font_size"])
    LOGGER.info("font applied: %s size=%s", config["font_name"], config["font_size"])

    alignment = config["alignment"]
    apply_global_alignment(ws, alignment["horizontal"], alignment["vertical"])
    LOGGER.info("alignment applied: %s + %s", alignment["horizontal"], alignment["vertical"])

    header_alignment = config["header_alignment"]
    apply_header_alignment(ws, header_alignment["horizontal"], header_alignment["vertical"])
    LOGGER.info(
        "header alignment override applied: %s + %s",
        header_alignment["horizontal"],
        header_alignment["vertical"],
    )

    result["layout_applied"] = True

    if config.get("autofit"):
        autofit_columns(ws)
        result["autofit_applied"] = True
        LOGGER.info("autofit applied")

    width_map = config.get("column_widths", {})
    applied_widths = apply_column_widths(ws, width_map) if width_map else {}
    result["custom_column_widths_applied"] = []
    if width_map:
        for header, count in applied_widths.items():
            result["custom_column_widths_applied"].extend([header] * count)
        LOGGER.info(
            "custom column widths applied: %s",
            ", ".join(result["custom_column_widths_applied"]) if result["custom_column_widths_applied"] else "(none)",
        )


def applyFormatting(ws, config: dict[str, Any], result: dict[str, Any]) -> None:
    apply_formatting(ws, config, result)


def verify_rows(ws, result: dict[str, Any]) -> list[dict[str, Any]]:
    min_row, min_col, max_row, _ = find_used_range(ws)
    candidate_rows = [min_row, min_row + (max_row - min_row) // 2, max_row]
    ordered_rows: list[int] = []
    for row_num in candidate_rows:
        if row_num not in ordered_rows:
            ordered_rows.append(row_num)
    verification: list[dict[str, Any]] = []
    for row_num in ordered_rows:
        cell = ws.cell(row=row_num, column=min_col)
        verification.append(
            {
                "row": row_num,
                "font_name": cell.font.name,
                "font_size": cell.font.sz,
                "horizontal": cell.alignment.horizontal,
                "vertical": cell.alignment.vertical,
                "cell": cell.coordinate,
            }
        )
    return verification


def scan_range_for_font_mismatches(ws, expected_font_name: str, expected_font_size: int) -> list[dict[str, Any]]:
    min_row, min_col, max_row, max_col = find_used_range(ws)
    mismatches: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            if float(cell.font.sz or 0) != float(expected_font_size):
                mismatches.append({"cell": cell.coordinate, "font_size": cell.font.sz})
            elif normalize_text(cell.font.name) != expected_font_name:
                mismatches.append({"cell": cell.coordinate, "font_name": cell.font.name})
    return mismatches


def verify_value_rows(ws) -> list[dict[str, Any]]:
    _, _, max_row, _ = find_used_range(ws)
    if max_row < 2:
        return []
    candidate_rows = [2, 2 + (max_row - 2) // 2, max_row]
    ordered_rows: list[int] = []
    for row_num in candidate_rows:
        if row_num not in ordered_rows:
            ordered_rows.append(row_num)
    verification = []
    for row_num in ordered_rows:
        source_value = normalize_text(ws.cell(row=row_num, column=2).value)
        match = re.search(r"(\d+)$", source_value)
        parsed_n = int(match.group(1)) if match else None
        expected_value = parsed_n // 10 if parsed_n is not None and 1000 <= parsed_n < 10000 else parsed_n
        cell_value = ws.cell(row=row_num, column=3).value
        verification.append(
            {
                "cell": f"C{row_num}",
                "source_cell": f"B{row_num}",
                "source_value": source_value,
                "parsed_n": parsed_n,
                "written_value": cell_value,
                "ok": cell_value == expected_value,
            }
        )
    return verification


def run_formatting(file_path: Path, file_type: str | None = None) -> dict[str, Any]:
    effective_file_type = file_type or DEFAULT_FILE_TYPE
    result = result_template(file_path=file_path, file_type=effective_file_type)

    if not file_path.exists():
        result["reason"] = "找不到目標檔案"
        result["existing_files"] = list_existing_files()
        return result

    detected_type = detect_file_type(file_path) if file_type in (None, DEFAULT_FILE_TYPE) else file_type
    result["detected_type"] = detected_type
    LOGGER.info("detected type = %s", detected_type or "unknown")

    if detected_type is None:
        result["reason"] = "無法判斷檔案類型"
        result["existing_files"] = list_existing_files()
        return result

    try:
        config = build_runtime_config(detected_type, file_path)
    except FormattingError as exc:
        result["reason"] = str(exc)
        return result

    result["applied_config"] = config["name"]
    LOGGER.info("applying config = %s", config["name"])
    result["file_found"] = True

    try:
        workbook = load_workbook(file_path)
    except PermissionError:
        result["reason"] = "儲存失敗"
        return result

    result["sheet_names"] = workbook.sheetnames
    sheet_name = config["sheet_name"]
    if sheet_name not in workbook.sheetnames:
        result["reason"] = "工作表不存在"
        workbook.close()
        return result

    ws = workbook[sheet_name]
    min_row, min_col, max_row, max_col = find_used_range(ws)
    result["used_range"] = {
        "min_row": min_row,
        "max_row": max_row,
        "min_col": min_col,
        "max_col": max_col,
    }
    result["headers"] = [
        normalize_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1), ())
    ]

    before_hook = config.get("before_format")
    if callable(before_hook):
        before_hook(ws, result, config, file_path)

    try:
        apply_formatting(ws, config, result)
    except FormattingError as exc:
        result["reason"] = str(exc)
        workbook.close()
        return result

    after_hook = config.get("after_format")
    if callable(after_hook):
        after_hook(ws, result, config, file_path)

    result["format_applied"] = True
    result["headers"] = [
        normalize_text(cell.value) for cell in next(ws.iter_rows(min_row=1, max_row=1), ())
    ]
    result["verification"] = verify_rows(ws, result)
    if detected_type == "fault_notice":
        result["value_verification"] = verify_value_rows(ws)

    try:
        workbook.save(file_path)
    except PermissionError:
        result["reason"] = "儲存失敗"
        workbook.close()
        return result
    workbook.close()
    result["saved"] = True

    verify_workbook = load_workbook(file_path)
    verify_ws = verify_workbook[sheet_name]
    verify_min_row, verify_min_col, verify_max_row, verify_max_col = find_used_range(verify_ws)
    result["used_range"] = {
        "min_row": verify_min_row,
        "max_row": verify_max_row,
        "min_col": verify_min_col,
        "max_col": verify_max_col,
    }
    result["headers"] = [
        normalize_text(cell.value) for cell in next(verify_ws.iter_rows(min_row=1, max_row=1), ())
    ]
    result["verification"] = verify_rows(verify_ws, result)
    result["header_bold_applied"] = bool(verify_ws["C1"].font.bold)
    if detected_type == "fault_notice":
        result["value_verification"] = verify_value_rows(verify_ws)

    mismatches = scan_range_for_font_mismatches(verify_ws, config["font_name"], config["font_size"])
    result["mismatch_cells"] = mismatches[:20]
    result["full_range_verified"] = len(mismatches) == 0
    verify_workbook.close()

    if mismatches:
        result["saved"] = False
        result["reason"] = "格式套用失敗"
        return result
    if detected_type == "fault_notice" and any(not item["ok"] for item in result["value_verification"]):
        result["saved"] = False
        result["reason"] = "車號寫入失敗"
        return result

    return result


def runFormatting(file_path: Path, file_type: str | None = None) -> dict[str, Any]:
    return run_formatting(file_path, file_type)


def format_fault_notice_excel(file_path: Path) -> dict[str, Any]:
    return run_formatting(file_path, "fault_notice")


def format_open_fault_notice_excel(file_path: Path) -> dict[str, Any]:
    return run_formatting(file_path, "open_fault_notice")


def find_latest_excel_file() -> Path | None:
    if not TARGET_DIR.exists():
        return None
    candidates = sorted(
        (
            path
            for path in TARGET_DIR.iterdir()
            if path.is_file() and path.suffix.lower() == ".xlsx" and not path.name.startswith("~$")
        ),
        key=lambda path: path.stat().st_mtime,
        reverse=True,
    )
    return candidates[0] if candidates else None


def get_patterns_for_file_type(file_type: str) -> tuple[str, ...]:
    pattern_map = {
        "fault_notice": FAULT_NOTICE_FILENAME_PATTERNS,
        "open_fault_notice": OPEN_FAULT_NOTICE_FILENAME_PATTERNS,
    }
    return pattern_map.get(file_type, ())


def resolve_target_file(*, file_path_arg: str | None, file_type: str) -> tuple[Path | None, str]:
    if file_path_arg:
        return Path(file_path_arg), "explicit_path"

    latest = find_latest_excel_file()
    if latest is None:
        return None, "latest_file"

    if file_type == DEFAULT_FILE_TYPE:
        return latest, "latest_file"

    patterns = get_patterns_for_file_type(file_type)
    matching = [
        path
        for path in TARGET_DIR.iterdir()
        if path.is_file()
        and path.suffix.lower() == ".xlsx"
        and not path.name.startswith("~$")
        and any(pattern in path.name for pattern in patterns)
    ]
    matching.sort(key=lambda path: path.stat().st_mtime, reverse=True)
    return (matching[0], "latest_matching_type") if matching else (latest, "latest_file")


def dispatch_formatter(file_path: Path, file_type: str) -> dict[str, Any]:
    effective_file_type = None if file_type == DEFAULT_FILE_TYPE else file_type
    return run_formatting(file_path, effective_file_type)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Format MMIS Excel files using shared config-driven engine")
    parser.add_argument("--file", dest="file_path", help="Explicit Excel file path")
    parser.add_argument(
        "--file-type",
        default=DEFAULT_FILE_TYPE,
        choices=["auto", "fault_notice", "open_fault_notice"],
        help="Formatter strategy to use; default is auto-detect from filename",
    )
    return parser


def main() -> int:
    args = build_parser().parse_args()
    file_path, selected_by = resolve_target_file(file_path_arg=args.file_path, file_type=args.file_type)

    if file_path is None:
        result = result_template(file_type=args.file_type)
        result["reason"] = "找不到目標檔案"
        result["existing_files"] = list_existing_files()
        result["selected_by"] = selected_by
        print(json.dumps(result, ensure_ascii=False))
        return 1

    result = dispatch_formatter(file_path, args.file_type)
    result["selected_by"] = selected_by

    if not result.get("file_found"):
        result["existing_files"] = list_existing_files()
        print(json.dumps(result, ensure_ascii=False))
        return 1

    if not result.get("saved"):
        print(json.dumps(result, ensure_ascii=False))
        return 1

    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    sys.exit(main())
