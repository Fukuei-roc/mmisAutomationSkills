from __future__ import annotations

import argparse
import contextlib
import json
import logging
import re
import sys
import time
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


SKILL_DIR = Path(__file__).resolve().parent.parent
MMIS_ONE_A_DIR = Path(
    r"C:\Users\NMMIS\.codex\skills\mmis-query-1a-work-order-linked-fault-notices\scripts"
)
MMIS_CORE_DIR = Path(r"C:\Users\NMMIS\.codex\skills\mmis-query-unprocessed-fault-notices\scripts")
for import_path in (MMIS_ONE_A_DIR, MMIS_CORE_DIR):
    if str(import_path) not in sys.path:
        sys.path.insert(0, str(import_path))

from playwright_linked_fault_notices_query import (  # noqa: E402
    DEFAULT_TIMEOUT_MS,
    DETAIL_WORK_ORDER_SELECTOR,
    LinkedFaultNoticeQuery,
    RESULT_ROW_SELECTOR,
    normalize_value,
)


TARGET_DIR = Path(
    r"C:\Users\NMMIS\OneDrive - Ministry of Transportation and Communications-7280502-Taiwan Railways Administration, MOTC\文件\MMIS桌面"
)
LOG_DIR = SKILL_DIR / "logs"
LOG_FILE = LOG_DIR / "mmis_auto_link_unprocessed_fault_notices.log"
DEBUG_SCREENSHOT_DIR = Path(r"C:\Users\NMMIS\Downloads\mmis_query_debug")
DEFAULT_DEPOT = "新竹機務段"
AUTOSAVE_EVERY = 5
MAX_ROW_RETRIES = 3
MAX_FIELD_RETRIES = 3
OUTPUT_COLUMN_LETTER = "I"
OUTPUT_HEADER = "日檢工單"
DEBUG_SCREENSHOT_VIEWPORT = {"width": 2400, "height": 1800}
DEBUG_SCREENSHOT_ZOOM = "0.75"
SKIPPABLE_OUTPUT_VALUES = {"找不到日檢單", "缺少查詢條件"}

QUERY_MENU_SELECTORS = [
    "#toolbar2_tbs_0_tbcb_0_query-img",
    "xpath=(//*[@id='toolbar2_tbs_0_tbcb_0_query-img' or contains(@id,'_query-img') or contains(@id,'query-img')])[1]",
]
ALL_RECORDS_OPTION_SELECTORS = [
    "#menu0_useAllRecsQuery_OPTION_a",
    "xpath=(//*[@id='menu0_useAllRecsQuery_OPTION_a' or contains(@id,'useAllRecsQuery_OPTION_a')])[1]",
    "xpath=(//*[self::a or self::span or self::div][normalize-space(.)='所有記錄'])[1]",
    "text=所有記錄",
]
C1_DEPOT_INPUT_SELECTORS = [
    "xpath=(//input[@id='m6a7dfd2f_tfrow_[C:1]_txt-tb' or (contains(@id,'tfrow_') and contains(@id,'[C:1]_txt-tb'))])[1]",
]
C2_CLEAR_INPUT_SELECTORS = [
    "xpath=(//input[@id='m6a7dfd2f_tfrow_[C:2]_txt-tb' or (contains(@id,'tfrow_') and contains(@id,'[C:2]_txt-tb'))])[1]",
]
CAR_NO_INPUT_SELECTORS = [
    "xpath=(//input[@id='m6a7dfd2f_tfrow_[C:3]_txt-tb' or (contains(@id,'tfrow_') and contains(@id,'[C:3]_txt-tb'))])[1]",
]
DATE_INPUT_SELECTORS = [
    "xpath=(//input[(contains(@id,'tfrow_') and contains(@id,'[C:11]_txt-tb'))])[1]",
    "xpath=(//input[@id='m6a7dfd2f_tfrow_[C:10]_txt-tb' or (contains(@id,'tfrow_') and contains(@id,'[C:10]_txt-tb'))])[1]",
    "xpath=(//input[(contains(@id,'tfrow_') and contains(@id,'[C:9]_txt-tb'))])[1]",
    "xpath=(//input[(contains(@id,'tfrow_') and contains(@id,'[C:12]_txt-tb'))])[1]",
]
FILTER_INPUTS_SELECTOR = "xpath=//input[contains(@id,'tfrow_') and contains(@id,'_txt-tb')]"
LOADING_OVERLAY_SELECTOR = "xpath=//*[@id='wait' or contains(@class,'bx--loading-overlay') or contains(@class,'wait')]"


@dataclass
class StepMetric:
    name: str
    status: str
    duration_ms: int
    details: dict[str, Any]


class AutoLinkError(RuntimeError):
    pass


def mmdd_today() -> str:
    return datetime.now().strftime("%m%d")


def build_target_filename() -> str:
    return f"本段未處理故障通報{mmdd_today()}.xlsx"


def build_logger() -> logging.Logger:
    LOG_DIR.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("mmis_auto_link_unprocessed_fault_notices")
    if logger.handlers:
        return logger

    logger.setLevel(logging.INFO)
    formatter = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s", datefmt="%Y-%m-%d %H:%M:%S")

    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")

    stream_handler = logging.StreamHandler(sys.stdout)
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    file_handler = logging.FileHandler(LOG_FILE, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)
    return logger


def normalize_header(value: Any) -> str:
    return normalize_value(str(value) if value is not None else "")


def safe_cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return normalize_value(str(value))


def get_column_index_map(ws: Worksheet) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for column_index in range(1, ws.max_column + 1):
        header = normalize_header(ws.cell(row=1, column=column_index).value)
        if header:
            mapping[header] = column_index
    return mapping


def format_excel_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y/%m/%d")
    if hasattr(value, "year") and hasattr(value, "month") and hasattr(value, "day"):
        return f"{value.year:04d}/{value.month:02d}/{value.day:02d}"

    raw = normalize_value(str(value))
    if not raw:
        return ""

    normalized = raw.replace("-", "/").replace(".", "/")
    for pattern in ("%Y/%m/%d", "%Y/%m/%d %H:%M:%S", "%Y/%m/%d %H:%M"):
        with contextlib.suppress(ValueError):
            parsed = datetime.strptime(normalized, pattern)
            return parsed.strftime("%Y/%m/%d")
    return raw


def formatQueryDate(date_value: Any) -> str:
    if date_value is None:
        return ""

    raw = normalize_value(str(date_value))
    if raw.startswith(">"):
        raw = normalize_value(raw[1:])

    normalized_date = format_excel_date(raw if raw else date_value)
    normalized_date = normalize_value(normalized_date)
    if not normalized_date:
        return ""
    return f">{normalized_date}"


def first_visible_locator(page, selectors: list[str]):
    last_error: Exception | None = None
    for selector in selectors:
        try:
            locator = page.locator(selector).first
            if locator.count() == 0:
                continue
            locator.wait_for(state="visible", timeout=5000)
            return locator
        except Exception as exc:  # noqa: BLE001
            last_error = exc
            continue
    raise AutoLinkError(f"找不到可用元素: {selectors}") from last_error


class AutoLinkUnprocessedFaultNotices(LinkedFaultNoticeQuery):
    def __init__(self, *, file_path: Path | None = None, skip_filled: bool = False) -> None:
        super().__init__()
        self.logger = build_logger()
        self.step_metrics: list[StepMetric] = []
        self.file_path = file_path
        self.skip_filled = skip_filled
        self.workbook = None
        self.worksheet: Worksheet | None = None
        self.column_map: dict[str, int] = {}
        self.output_column_index = 9
        self._all_records_selected = False
        self.success_count = 0
        self.fail_count = 0
        self.processed_count = 0
        self._last_query_signature: tuple[str, str] | None = None
        self._last_result_signature: str | None = None

    @contextlib.contextmanager
    def timed_step(self, name: str, **details: Any):
        started_at = time.perf_counter()
        self.logger.info("START %s%s", name, self._format_details(details))
        status = "success"
        try:
            yield
        except Exception:
            status = "failed"
            raise
        finally:
            duration_ms = int((time.perf_counter() - started_at) * 1000)
            self.step_metrics.append(StepMetric(name=name, status=status, duration_ms=duration_ms, details=details))
            level = logging.INFO if status == "success" else logging.ERROR
            self.logger.log(level, "END %s status=%s duration_ms=%s", name, status, duration_ms)

    def resolve_target_file(self) -> Path:
        if self.file_path is not None:
            if not self.file_path.exists():
                raise AutoLinkError("找不到指定 Excel 檔案")
            return self.file_path

        target = TARGET_DIR / build_target_filename()
        if not target.exists():
            raise AutoLinkError("找不到當日未處理故障通報檔案")
        return target

    def load_excel(self) -> Path:
        with self.timed_step("load_excel"):
            target = self.resolve_target_file()
            self.workbook = load_workbook(target)
            self.worksheet = self.workbook.active
            assert self.worksheet is not None
            self.column_map = get_column_index_map(self.worksheet)
            if "發生日期" not in self.column_map or "車號" not in self.column_map:
                raise AutoLinkError("Excel 缺少必要欄位：發生日期或車號")
            self.output_column_index = self._ensure_output_column()
            self.file_path = target
            total_rows = max(self.worksheet.max_row - 1, 0)
            self.logger.info("[INFO] file loaded")
            self.logger.info("[INFO] total rows: %s", total_rows)
            return target

    def _ensure_output_column(self) -> int:
        assert self.worksheet is not None
        column_index = self.worksheet[OUTPUT_COLUMN_LETTER + "1"].column
        header_cell = self.worksheet[f"{OUTPUT_COLUMN_LETTER}1"]
        if not normalize_header(header_cell.value):
            header_cell.value = OUTPUT_HEADER
        return column_index

    def should_skip_row(self, output_value: Any) -> bool:
        normalized = normalize_value(str(output_value) if output_value is not None else "")
        if not normalized:
            return False
        if self.skip_filled:
            return True
        if normalized in SKIPPABLE_OUTPUT_VALUES:
            return True
        return bool(re.fullmatch(r"\d{3}-1A-\d+", normalized))

    def select_all_records_mode(self) -> None:
        def action() -> None:
            with self.timed_step("select_all_records_mode"):
                assert self.page is not None
                query_button = first_visible_locator(self.page, QUERY_MENU_SELECTORS)
                query_button.click()
                self.page.get_by_text("所有記錄", exact=True).first.wait_for(
                    state="visible", timeout=DEFAULT_TIMEOUT_MS
                )
                option = first_visible_locator(self.page, ALL_RECORDS_OPTION_SELECTORS)
                option.click()
                self.wait_for_loading_complete()
                first_visible_locator(self.page, C1_DEPOT_INPUT_SELECTORS).wait_for(
                    state="visible", timeout=DEFAULT_TIMEOUT_MS
                )
                self._all_records_selected = True
                self.logger.info("[DEBUG] switched to all records = true")

        self._with_retry("select_all_records_mode", action)

    def _clear_filter_inputs(self) -> None:
        assert self.page is not None
        inputs = self.page.locator(FILTER_INPUTS_SELECTOR)
        for index in range(inputs.count()):
            locator = inputs.nth(index)
            with contextlib.suppress(Exception):
                locator.fill("")

    def wait_for_loading_complete(self, timeout_ms: int = DEFAULT_TIMEOUT_MS) -> None:
        assert self.page is not None
        with contextlib.suppress(Exception):
            self.page.wait_for_load_state("networkidle")
        with contextlib.suppress(Exception):
            self.page.locator(LOADING_OVERLAY_SELECTOR).last.wait_for(state="hidden", timeout=timeout_ms)
        self.page.wait_for_timeout(120)

    def clear_input_by_keyboard(self, locator) -> None:
        assert self.page is not None
        for _ in range(MAX_FIELD_RETRIES):
            locator.click()
            locator.press("Control+A")
            locator.press("Backspace")
            if normalize_value(locator.input_value()):
                locator.press("Delete")
            if not normalize_value(locator.input_value()):
                return
            self.page.wait_for_timeout(120)
        remaining = normalize_value(locator.input_value())
        raise AutoLinkError(f"欄位清除失敗，仍有值: {remaining}")

    def fill_input_with_verification(self, locator, expected_value: str, *, field_name: str) -> None:
        assert self.page is not None
        for _ in range(MAX_FIELD_RETRIES):
            locator.click()
            locator.press("Control+A")
            locator.press("Backspace")
            locator.fill(expected_value)
            actual_value = normalize_value(locator.input_value())
            if actual_value == expected_value:
                return
            self.page.wait_for_timeout(120)
        raise AutoLinkError(
            f"{field_name} 欄位寫入失敗: expected={expected_value}, actual={normalize_value(locator.input_value())}"
        )

    def _return_to_list_page(self) -> None:
        assert self.page is not None
        self.page.go_back(wait_until="domcontentloaded")
        self.wait_for_loading_complete(timeout_ms=5000)
        first_visible_locator(self.page, C1_DEPOT_INPUT_SELECTORS).wait_for(state="visible", timeout=5000)

    def _has_no_results(self) -> bool:
        assert self.page is not None
        no_result_patterns = ["查無資料", "沒有資料", "沒有要顯示的列。", "No records to display"]
        body_text = self.page.locator("body").inner_text(timeout=5000)
        return any(pattern in body_text for pattern in no_result_patterns) and self.page.locator(RESULT_ROW_SELECTOR).count() == 0

    def _first_result_signature(self) -> str:
        assert self.page is not None
        with contextlib.suppress(Exception):
            locator = self.page.locator(RESULT_ROW_SELECTOR).first
            if locator.count() == 0:
                return ""
            return normalize_value(locator.inner_text(timeout=3000))
        return ""

    def wait_for_search_result_state(
        self,
        *,
        previous_result_signature: str,
        query_signature: tuple[str, str],
        timeout_ms: int = 15000,
    ) -> str:
        assert self.page is not None
        started_at = time.perf_counter()
        while (time.perf_counter() - started_at) * 1000 < timeout_ms:
            self.wait_for_loading_complete(timeout_ms=3000)
            if self._has_no_results():
                return "no_result"
            current_signature = self._first_result_signature()
            if current_signature:
                if query_signature == self._last_query_signature and current_signature == self._last_result_signature:
                    return "has_result"
                if current_signature != previous_result_signature:
                    return "has_result"
            self.page.wait_for_timeout(250)
        return "timeout"

    def save_query_debug_screenshot(self, *, row_number: int) -> None:
        assert self.page is not None
        DEBUG_SCREENSHOT_DIR.mkdir(parents=True, exist_ok=True)
        target = DEBUG_SCREENSHOT_DIR / f"query_debug_{row_number}.png"
        try:
            self.page.set_viewport_size(DEBUG_SCREENSHOT_VIEWPORT)
            previous_zoom = self.page.evaluate(
                "() => document.body ? (document.body.style.zoom || '') : ''"
            )
            self.page.evaluate(
                f"""() => {{
                    if (document.body) {{
                        document.body.style.zoom = "{DEBUG_SCREENSHOT_ZOOM}";
                        window.scrollTo(0, 0);
                    }}
                }}"""
            )
            self.page.screenshot(path=str(target), full_page=True)
            self.page.evaluate(
                """(previousZoom) => {
                    if (document.body) {
                        document.body.style.zoom = previousZoom || '';
                    }
                }""",
                previous_zoom,
            )
            self.logger.info("[DEBUG] screenshot saved: %s", target)
        except Exception as exc:  # noqa: BLE001
            self.logger.error("failed to save debug screenshot for row %s: %s", row_number, exc)

    def search_daily_check_work_order(self, *, row_number: int, date_query: str, car_no: str) -> str | None:
        def action() -> str | None:
            with self.timed_step(
                "search_daily_check_work_order",
                row_number=row_number,
                date_query=date_query,
                car_no=car_no,
            ):
                self.open1AWorkOrderPage()
                self.select_all_records_mode()
                assert self.page is not None

                c1_input = first_visible_locator(self.page, C1_DEPOT_INPUT_SELECTORS)
                c2_input = first_visible_locator(self.page, C2_CLEAR_INPUT_SELECTORS)
                date_input = first_visible_locator(self.page, DATE_INPUT_SELECTORS)
                car_input = first_visible_locator(self.page, CAR_NO_INPUT_SELECTORS)

                self._clear_filter_inputs()
                self.fill_input_with_verification(c1_input, DEFAULT_DEPOT, field_name="C1")
                self.clear_input_by_keyboard(c2_input)
                self.fill_input_with_verification(date_input, date_query, field_name="C11")
                self.fill_input_with_verification(car_input, car_no, field_name="C3")
                c1_value = normalize_value(c1_input.input_value())
                c2_value = normalize_value(c2_input.input_value())
                c3_value = normalize_value(car_input.input_value())
                c11_value = normalize_value(date_input.input_value())
                self.logger.info("[DEBUG] C1 value = %s", c1_value)
                self.logger.info("[DEBUG] C2 value = %s", c2_value)
                self.logger.info("[DEBUG] C3 value = %s", c3_value)
                self.logger.info("[DEBUG] C11 value = %s", c11_value)
                if c1_value != DEFAULT_DEPOT or c2_value != "" or c3_value != car_no or c11_value != date_query:
                    raise AutoLinkError(
                        f"查詢前欄位驗證失敗: C1={c1_value}, C2={c2_value}, C3={c3_value}, C11={c11_value}"
                    )
                self.logger.info("[DEBUG] query values: 日期=%s 車號=%s", date_query, car_no)
                self.save_query_debug_screenshot(row_number=row_number)
                previous_result_signature = self._first_result_signature()
                query_signature = (date_query, car_no)
                car_input.click()
                self.page.wait_for_timeout(150)
                if not car_input.evaluate("(element) => document.activeElement === element"):
                    raise AutoLinkError("車號欄位未取得 focus，無法穩定觸發查詢")
                self.logger.info("[DEBUG] focus set to C3 (車號欄位)")
                self.logger.info("[DEBUG] press Enter to trigger query")
                car_input.press("Enter")
                search_state = self.wait_for_search_result_state(
                    previous_result_signature=previous_result_signature,
                    query_signature=query_signature,
                )

                if search_state == "no_result":
                    self.logger.info("[INFO] no result")
                    self._last_query_signature = query_signature
                    self._last_result_signature = ""
                    return None
                if search_state == "timeout":
                    raise AutoLinkError("查詢後結果狀態未穩定")

                first_result = self.page.locator(RESULT_ROW_SELECTOR).first
                if first_result.count() == 0:
                    self.logger.info("[INFO] no result")
                    return None

                first_result.wait_for(state="visible", timeout=DEFAULT_TIMEOUT_MS)
                self.wait_for_loading_complete()
                for _ in range(MAX_FIELD_RETRIES):
                    try:
                        first_result.click(timeout=5000)
                        break
                    except Exception:  # noqa: BLE001
                        self.wait_for_loading_complete(timeout_ms=5000)
                else:
                    first_result.click(timeout=DEFAULT_TIMEOUT_MS)
                detail_input = self.page.locator(DETAIL_WORK_ORDER_SELECTOR).first
                detail_input.wait_for(state="visible", timeout=DEFAULT_TIMEOUT_MS)
                work_order_no = normalize_value(detail_input.input_value())
                if not work_order_no:
                    raise AutoLinkError("進入明細後讀不到工單號")
                self.logger.info("[INFO] found work order: %s", work_order_no)
                self._last_query_signature = query_signature
                self._last_result_signature = self._first_result_signature() or work_order_no
                try:
                    self._return_to_list_page()
                except Exception as exc:  # noqa: BLE001
                    self.logger.warning("return to list page failed after finding work order %s: %s", work_order_no, exc)
                    self._on_1a_page = False
                return work_order_no

        return self._with_retry("search_daily_check_work_order", action, retries=MAX_ROW_RETRIES)

    def process_rows(self) -> dict[str, Any]:
        started_at = time.perf_counter()
        target = self.load_excel()
        assert self.worksheet is not None

        self.ensureLoggedIn()
        self.open1AWorkOrderPage()
        self.select_all_records_mode()

        date_column = self.column_map["發生日期"]
        car_column = self.column_map["車號"]

        for row_index in range(2, self.worksheet.max_row + 1):
            self.processed_count += 1
            self.logger.info("[INFO] processing row %s", row_index - 1)
            output_cell = self.worksheet.cell(row=row_index, column=self.output_column_index)
            if self.should_skip_row(output_cell.value):
                self.logger.info("[INFO] skip filled row %s", row_index - 1)
                continue

            date_value = self.worksheet.cell(row=row_index, column=date_column).value
            car_no = safe_cell_text(self.worksheet.cell(row=row_index, column=car_column).value)
            formatted_date = format_excel_date(date_value)
            query_date = formatQueryDate(date_value)
            self.logger.info("[DEBUG] raw excel date = %s", formatted_date)
            self.logger.info("[DEBUG] formatted query date = %s", query_date)
            if not formatted_date or not query_date or not car_no:
                output_cell.value = "缺少查詢條件"
                self.fail_count += 1
                self.save_workbook()
                continue

            self.logger.info("[INFO] searching: 日期=%s 車號=%s", formatted_date, car_no)

            try:
                work_order_no = self.search_daily_check_work_order(
                    row_number=row_index - 1,
                    date_query=query_date,
                    car_no=car_no,
                )
                if work_order_no:
                    output_cell.value = work_order_no
                    self.success_count += 1
                else:
                    output_cell.value = "找不到日檢單"
                    self.fail_count += 1
            except Exception as exc:  # noqa: BLE001
                self.logger.error("row %s failed: %s", row_index, exc)
                output_cell.value = "查詢失敗"
                self.fail_count += 1
                self.save_workbook()
                self._all_records_selected = False
                self._on_1a_page = False
                self.close()
                continue

            self.save_workbook()

        self.save_workbook()
        elapsed_ms = int((time.perf_counter() - started_at) * 1000)
        self.logger.info("[INFO] completed")
        self.logger.info("[INFO] success count: %s", self.success_count)
        self.logger.info("[INFO] fail count: %s", self.fail_count)
        return {
            "ok": True,
            "file_path": str(target),
            "total_rows": max(self.worksheet.max_row - 1, 0),
            "success_count": self.success_count,
            "fail_count": self.fail_count,
            "log_file": str(LOG_FILE),
            "elapsed_ms": elapsed_ms,
            "step_metrics": [asdict(metric) for metric in self.step_metrics],
        }

    def save_workbook(self) -> None:
        assert self.workbook is not None
        assert self.file_path is not None
        last_error: Exception | None = None
        for _ in range(MAX_FIELD_RETRIES):
            try:
                self.workbook.save(self.file_path)
                return
            except PermissionError as exc:
                last_error = exc
                time.sleep(0.5)
        if last_error is not None:
            raise last_error


def run_auto_link(*, file_path: str | None = None, skip_filled: bool = False) -> dict[str, Any]:
    linker = AutoLinkUnprocessedFaultNotices(
        file_path=Path(file_path) if file_path else None,
        skip_filled=skip_filled,
    )
    try:
        return linker.process_rows()
    except AutoLinkError as exc:
        return {"ok": False, "error": str(exc), "log_file": str(LOG_FILE)}
    finally:
        linker.close()


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Auto-link today's unprocessed fault notices to MMIS 1A work orders")
    parser.add_argument("--file", help="Optional Excel file path override")
    parser.add_argument("--skip-filled", action="store_true", help="Skip rows whose output column I already has data")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    result = run_auto_link(file_path=args.file, skip_filled=args.skip_filled)
    print(json.dumps(result, ensure_ascii=False))
    return 0 if result.get("ok") else 1


if __name__ == "__main__":
    raise SystemExit(main())
